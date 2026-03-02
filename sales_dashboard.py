"""
LNDR Sales Dashboard
Streamlit + BigQuery — live sales data, current year vs prior year.

Local:  streamlit run sales_dashboard.py
Cloud:  deploy to share.streamlit.io — set [gcp_service_account] secret.
"""

import calendar
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from google.cloud import bigquery
from google.oauth2 import service_account

# ── Config ────────────────────────────────────────────────────────────────────

BQ_PROJECT = "lndr-brain"
PROMO_XLSX = Path(__file__).parent / "promo_calendar.xlsx"

ALL_REGIONS = ["AU", "US", "UK", "EU", "ROW"]

EU_COUNTRIES = [
    "Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus", "Czech Republic",
    "Denmark", "Estonia", "Finland", "France", "Germany", "Greece", "Hungary",
    "Ireland", "Italy", "Latvia", "Lithuania", "Luxembourg", "Malta",
    "Netherlands", "Poland", "Portugal", "Romania", "Slovakia", "Slovenia",
    "Spain", "Sweden",
]
_EU_IN = "'" + "','".join(EU_COUNTRIES) + "'"

REGION_CASE = f"""CASE
  WHEN o.store = 'AU' THEN 'AU'
  WHEN o.store = 'US' THEN 'US'
  WHEN o.store = 'ROW' AND o.shipping_country = 'United Kingdom' THEN 'UK'
  WHEN o.store = 'ROW' AND o.shipping_country IN ({_EU_IN}) THEN 'EU'
  ELSE 'ROW'
END"""

SALE_DATE_CASE = """CASE o.store
  WHEN 'AU'  THEN DATE(DATETIME(o.created_at, 'Australia/Sydney'))
  WHEN 'US'  THEN DATE(DATETIME(o.created_at, 'America/New_York'))
  WHEN 'ROW' THEN DATE(DATETIME(o.created_at, 'Europe/London'))
END"""

COLORS = {
    "cur":  "#1a1a3e",
    "pri":  "#999999",
    "req":  "#e63946",
    "AU":   "#2196F3",
    "US":   "#4CAF50",
    "UK":   "#9C27B0",
    "EU":   "#FF9800",
    "ROW":  "#607D8B",
}

MONTH_NAMES = [date(2000, m, 1).strftime("%B") for m in range(1, 13)]
MONTH_SHORT  = [date(2000, m, 1).strftime("%b")  for m in range(1, 13)]

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="LNDR Sales",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── BigQuery client ───────────────────────────────────────────────────────────

@st.cache_resource
def get_bq():
    if "gcp_service_account" in st.secrets:
        creds = service_account.Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/bigquery"],
        )
        return bigquery.Client(credentials=creds, project=BQ_PROJECT)
    else:
        # Local dev — use Application Default Credentials (gcloud auth)
        try:
            return bigquery.Client(project=BQ_PROJECT)
        except Exception:
            st.error(
                "**BigQuery credentials not found.**\n\n"
                "On Streamlit Cloud: add `[gcp_service_account]` in **Settings → Secrets**.\n\n"
                "Locally: run `gcloud auth application-default login`."
            )
            st.stop()


# ── SQL helpers ───────────────────────────────────────────────────────────────

from datetime import timedelta

def _build_query(utc_ranges: list, tz_ranges: list) -> str:
    """
    Build a CTE-based query using o.total_price (matches weekly email exactly),
    proportionally allocated to line items for product-type filtering.
    Stage 1: UTC pre-filter for partition pruning.
    Stage 2: Timezone-accurate sale_date filter.
    """
    utc_clauses = " OR ".join(
        f"DATE(o.created_at) BETWEEN '{s}' AND '{e}'" for s, e in utc_ranges
    )
    tz_clauses = " OR ".join(f"""(
    (store = 'AU'  AND sale_date BETWEEN '{s}' AND '{e}')
 OR (store = 'US'  AND sale_date BETWEEN '{s}' AND '{e}')
 OR (store = 'ROW' AND sale_date BETWEEN '{s}' AND '{e}')
    )""" for s, e in tz_ranges)

    return f"""
    WITH pre AS (
      SELECT
        o.store,
        o.shipping_country,
        CASE o.store
          WHEN 'AU'  THEN DATE(DATETIME(o.created_at, 'Australia/Sydney'))
          WHEN 'US'  THEN DATE(DATETIME(o.created_at, 'America/New_York'))
          WHEN 'ROW' THEN DATE(DATETIME(o.created_at, 'Europe/London'))
        END AS sale_date,
        {REGION_CASE} AS region,
        COALESCE(NULLIF(TRIM(p.product_type), ''), 'Other') AS product_type,
        -- Order total in GBP (same as weekly email)
        o.total_price * CASE o.store
          WHEN 'AU'  THEN 1.0 / fx.aud
          WHEN 'US'  THEN 1.0 / fx.usd
          WHEN 'ROW' THEN 1.0
        END AS order_rev_gbp,
        -- Proportional allocation by line-item gross
        li.price * li.quantity AS line_gross,
        SUM(li.price * li.quantity) OVER (PARTITION BY o.order_id, o.store) AS order_gross
      FROM `lndr-brain.shopify_raw.orders` o
      JOIN `lndr-brain.shopify_raw.order_line_items` li
        ON o.order_id = li.order_id AND o.store = li.store
      JOIN `lndr-brain.reference.exchange_rates` fx
        ON fx.date = DATE(o.created_at)
      LEFT JOIN (
        SELECT product_id, store, MAX(product_type) AS product_type
        FROM `lndr-brain.shopify_raw.products`
        GROUP BY product_id, store
      ) p ON li.product_id = p.product_id AND li.store = p.store
      WHERE o.financial_status != 'voided'
        AND NOT REGEXP_CONTAINS(LOWER(IFNULL(o.tags, '')), r'swap')
        AND NOT (o.source_name = 'shopify_draft_order' AND o.total_price = 0)
        AND ({utc_clauses})
    )
    SELECT
      EXTRACT(DAY   FROM sale_date) AS day,
      EXTRACT(MONTH FROM sale_date) AS month_num,
      EXTRACT(YEAR  FROM sale_date) AS year_label,
      region,
      product_type,
      ROUND(SUM(order_rev_gbp * SAFE_DIVIDE(line_gross, order_gross))) AS revenue_gbp
    FROM pre
    WHERE {tz_clauses}
    GROUP BY day, month_num, year_label, region, product_type
    ORDER BY year_label, month_num, day
    """


def _utc_buffered(d_start: date, d_end: date) -> tuple:
    """Add 2-day buffer for UTC pre-filter (accounts for timezone offsets up to UTC+14)."""
    return (d_start - timedelta(days=2), d_end + timedelta(days=1))


# ── Data loaders (cached) ─────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner="Querying BigQuery...")
def load_monthly_data(year: int, month: int) -> pd.DataFrame:
    """
    Daily revenue by (day, region, product_type) for year/month AND prior year/month.
    Columns: day, month_num, year_label, region, product_type, revenue_gbp
    """
    prior = year - 1
    cur_s = date(year,  month, 1)
    cur_e = date(year,  month, calendar.monthrange(year,  month)[1])
    pri_s = date(prior, month, 1)
    pri_e = date(prior, month, calendar.monthrange(prior, month)[1])

    utc_ranges = [_utc_buffered(cur_s, cur_e), _utc_buffered(pri_s, pri_e)]
    tz_ranges  = [(cur_s, cur_e), (pri_s, pri_e)]

    df = get_bq().query(_build_query(utc_ranges, tz_ranges)).to_dataframe()
    df["day"]        = df["day"].astype(int)
    df["year_label"] = df["year_label"].astype(int)
    return df


@st.cache_data(ttl=1800, show_spinner="Querying BigQuery...")
def load_annual_data(year: int) -> pd.DataFrame:
    """
    Monthly revenue by (month_num, region, product_type) for year AND prior year.
    Columns: day, month_num, year_label, region, product_type, revenue_gbp
    """
    prior = year - 1
    cur_s = date(year,  1,  1)
    cur_e = date(year,  12, 31)
    pri_s = date(prior, 1,  1)
    pri_e = date(prior, 12, 31)

    utc_ranges = [_utc_buffered(cur_s, cur_e), _utc_buffered(pri_s, pri_e)]
    tz_ranges  = [(cur_s, cur_e), (pri_s, pri_e)]

    df = get_bq().query(_build_query(utc_ranges, tz_ranges)).to_dataframe()
    df["month_num"]  = df["month_num"].astype(int)
    df["year_label"] = df["year_label"].astype(int)
    return df


@st.cache_data(ttl=3600)
def load_product_types() -> list:
    bq = get_bq()
    sql = """
    SELECT DISTINCT COALESCE(NULLIF(TRIM(product_type), ''), 'Other') AS pt
    FROM `lndr-brain.shopify_raw.products`
    WHERE product_type IS NOT NULL
    ORDER BY pt
    """
    return [r.pt for r in bq.query(sql).result()]


@st.cache_data(ttl=1800)
def load_forecast(year: int, month: int) -> float:
    bq = get_bq()
    sql = f"""
    SELECT SUM(total_gbp) AS total
    FROM `lndr-brain.reference.monthly_forecast`
    WHERE forecast_year = {year} AND forecast_month = {month}
    """
    rows = list(bq.query(sql).result())
    return float(rows[0].total or 0) if rows and rows[0].total else 0.0


@st.cache_data(ttl=86400)
def load_promos_for_month(year: int, month: int) -> list:
    """Returns [(day, label), ...] for promo START days in year/month."""
    try:
        wb = openpyxl.load_workbook(str(PROMO_XLSX), read_only=True, data_only=True)
        ws = wb.worksheets[0]
    except Exception:
        return []
    first_day = {}
    for row in ws.iter_rows(values_only=True):
        d, campaign = row[2], row[3]
        if not isinstance(d, datetime) or not campaign:
            continue
        if d.year != year or d.month != month:
            continue
        key = str(campaign).strip()
        if key not in first_day or d.day < first_day[key]:
            first_day[key] = d.day
    return sorted([(day, label[:25]) for label, day in first_day.items()])


@st.cache_data(ttl=86400)
def load_promo_months(year: int) -> set:
    """Returns set of month numbers that had any promo in year."""
    try:
        wb = openpyxl.load_workbook(str(PROMO_XLSX), read_only=True, data_only=True)
        ws = wb.worksheets[0]
    except Exception:
        return set()
    months = set()
    for row in ws.iter_rows(values_only=True):
        d, campaign = row[2], row[3]
        if isinstance(d, datetime) and d.year == year and campaign:
            months.add(d.month)
    return months


# ── Formatting ────────────────────────────────────────────────────────────────

def gbp(v: float) -> str:
    if abs(v) >= 1_000_000:
        return f"£{v/1_000_000:.2f}M"
    if abs(v) >= 1_000:
        return f"£{v/1_000:.0f}k"
    return f"£{v:,.0f}"


def pct(v: float) -> str:
    return f"{v:+.1f}%"


def _chart_layout(title: str, xaxis: dict, height: int = 460) -> dict:
    return dict(
        title=dict(text=title, font=dict(size=14)),
        xaxis=dict(**xaxis),
        yaxis=dict(tickformat="£,.0f", gridcolor="#2a2a3e"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        hovermode="x unified",
        height=height,
        margin=dict(t=60, b=40, l=10, r=10),
        plot_bgcolor="#0f1117",
        paper_bgcolor="#0f1117",
        font=dict(color="#ccc"),
    )


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 📊 LNDR Sales")

    if st.button("🔄 Refresh", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.divider()

    view = st.radio("View", ["Monthly", "12-Month"], horizontal=True)

    today  = date.today()
    year   = st.selectbox("Year", list(range(today.year, today.year - 4, -1)))
    prior  = year - 1

    if view == "Monthly":
        default_m = (today.month - 1) if year == today.year else 0
        month_name = st.selectbox("Month", MONTH_NAMES, index=default_m)
        month = MONTH_NAMES.index(month_name) + 1
    else:
        month = None

    st.divider()

    regions = st.multiselect("Regions", ALL_REGIONS, default=ALL_REGIONS)
    if not regions:
        regions = ALL_REGIONS

    all_cats = load_product_types()
    cats = st.multiselect("Categories", all_cats, default=all_cats)
    if not cats:
        cats = all_cats

    st.divider()
    show_promos = st.toggle("Show promos", value=True)

    st.caption("Revenue uses line-item prices. "
               "Minor variance vs order totals may occur due to order-level discounts.")


# ── Load & filter data ────────────────────────────────────────────────────────

if view == "Monthly":
    raw = load_monthly_data(year, month)
else:
    raw = load_annual_data(year)

df = raw[raw["region"].isin(regions) & raw["product_type"].isin(cats)].copy()

cur_df_all = df[df["year_label"] == year]
pri_df_all = df[df["year_label"] == prior]


# ── MONTHLY VIEW ─────────────────────────────────────────────────────────────

if view == "Monthly":
    days_in_month  = calendar.monthrange(year, month)[1]
    is_current     = (year == today.year and month == today.month)

    cur_by_day = cur_df_all.groupby("day")["revenue_gbp"].sum()
    pri_by_day = pri_df_all.groupby("day")["revenue_gbp"].sum()

    last_data_day  = int(cur_by_day.index.max()) if not cur_by_day.empty else 0
    days_remaining = days_in_month - last_data_day

    cur_total = float(cur_by_day.sum())
    # Prior year apples-to-apples: only days we have current year data for
    pri_comparable = float(pri_by_day[pri_by_day.index <= last_data_day].sum())
    yoy_pct = (cur_total - pri_comparable) / pri_comparable * 100 if pri_comparable else 0

    forecast    = load_forecast(year, month)
    needed      = forecast - cur_total
    req_daily   = needed / days_remaining if days_remaining > 0 and needed > 0 else 0

    # ── Title + KPI cards ─────────────────────────────────────────────────────
    st.title(f"Global Gross Revenue — {date(year, month, 1).strftime('%B %Y')}")

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("MTD Revenue",     gbp(cur_total),
              delta=pct(yoy_pct) + " YoY",
              delta_color="normal")
    k2.metric("vs Forecast",
              f"{cur_total/forecast*100:.0f}%" if forecast else "—",
              delta=f"target {gbp(forecast)}",
              delta_color="off")
    k3.metric("Prior Year (same days)", gbp(pri_comparable))
    if is_current and req_daily > 0:
        k4.metric("Required daily avg",   gbp(req_daily),
                  delta=f"{days_remaining} days remaining",
                  delta_color="off")
    else:
        pri_full = float(pri_by_day.sum())
        k4.metric("Full month YoY",
                  pct((cur_total - pri_full) / pri_full * 100) if pri_full else "—")

    # ── Main line chart ───────────────────────────────────────────────────────
    fig = go.Figure()

    # Remaining-days shading
    if is_current and days_remaining > 0:
        fig.add_vrect(
            x0=last_data_day + 0.5, x1=days_in_month + 0.5,
            fillcolor="steelblue", opacity=0.07,
            layer="below", line_width=0,
        )

    # Prior year line
    pri_x = sorted(pri_by_day.index.tolist())
    pri_y = [float(pri_by_day.get(d, 0)) for d in pri_x]
    fig.add_trace(go.Scatter(
        x=pri_x, y=pri_y,
        mode="lines+markers",
        name=f"{date(prior, month, 1).strftime('%b %Y')}",
        line=dict(color=COLORS["pri"], dash="dash", width=1.5),
        marker=dict(size=3),
        customdata=[gbp(v) for v in pri_y],
        hovertemplate="%{customdata}<extra></extra>",
    ))

    # Current year line
    cur_x = sorted(cur_by_day.index.tolist())
    cur_y = [float(cur_by_day.get(d, 0)) for d in cur_x]
    fig.add_trace(go.Scatter(
        x=cur_x, y=cur_y,
        mode="lines+markers",
        name=f"{date(year, month, 1).strftime('%b %Y')} (to {last_data_day})",
        line=dict(color=COLORS["cur"], width=2),
        marker=dict(size=3),
        customdata=[gbp(v) for v in cur_y],
        hovertemplate="%{customdata}<extra></extra>",
    ))

    # Required daily avg line
    if is_current and req_daily > 0:
        req_x = list(range(last_data_day, days_in_month + 1))
        fig.add_trace(go.Scatter(
            x=req_x, y=[req_daily] * len(req_x),
            mode="lines+markers",
            name=f"Required avg ({gbp(req_daily)}/day)",
            line=dict(color=COLORS["req"], dash="dot", width=2),
            marker=dict(symbol="square", size=4),
            customdata=[gbp(req_daily)] * len(req_x),
            hovertemplate="%{customdata}<extra></extra>",
        ))
        mid_x = last_data_day + max(2, days_remaining * 2 // 3)
        all_y = pri_y + cur_y
        y_bump = (max(all_y) - min(all_y)) * 0.18 if all_y else req_daily * 0.2
        fig.add_annotation(
            x=min(mid_x, days_in_month - 1), y=req_daily + y_bump,
            text=f"Need {gbp(req_daily)}/day<br>({days_remaining} days left)<br>"
                 f"to hit {gbp(forecast)} target",
            showarrow=True, ax=0, ay=-40,
            font=dict(color=COLORS["req"], size=10),
            bgcolor="rgba(15,17,23,0.9)",
            bordercolor=COLORS["req"],
            borderwidth=1,
        )
        fig.add_annotation(
            x=(last_data_day + days_in_month) / 2,
            y=0, yref="paper",
            text="Remaining days", showarrow=False,
            font=dict(color="steelblue", size=10),
            yanchor="bottom",
        )

    # Promo annotations
    if show_promos:
        for i, (day, label) in enumerate(load_promos_for_month(prior, month)):
            y_val = float(pri_by_day.get(day, 0))
            if y_val > 0:
                fig.add_annotation(
                    x=day, y=y_val,
                    text=label, showarrow=True,
                    ax=0, ay=-(35 + 18 * (i % 3)),
                    font=dict(color="#888", size=9),
                    arrowcolor="#aaa", arrowwidth=1,
                    bgcolor="rgba(15,17,23,0.85)",
                )

        for i, (day, label) in enumerate(load_promos_for_month(year, month)):
            y_val = float(cur_by_day.get(day, 0))
            if y_val > 0:
                fig.add_annotation(
                    x=day, y=y_val,
                    text=label, showarrow=True,
                    ax=0, ay=-(35 + 18 * (i % 3)),
                    font=dict(color=COLORS["cur"], size=9),
                    arrowcolor=COLORS["cur"], arrowwidth=1,
                    bgcolor="rgba(15,17,23,0.85)",
                )

    fig.update_layout(**_chart_layout(
        f"Daily Revenue — {date(year, month, 1).strftime('%b %Y')} vs "
        f"{date(prior, month, 1).strftime('%b %Y')}",
        xaxis=dict(
            tickmode="linear", tick0=1, dtick=1,
            range=[0.5, days_in_month + 0.5],
            gridcolor="#2a2a3e",
        ),
    ))
    st.plotly_chart(fig, use_container_width=True)

    # ── Secondary charts ──────────────────────────────────────────────────────
    col_l, col_r = st.columns(2)

    with col_l:
        st.subheader("By Region")
        reg_cur = cur_df_all.groupby("region")["revenue_gbp"].sum().reset_index()
        # Only include days up to last_data_day for prior year comparability
        reg_pri_comp = (
            pri_df_all[pri_df_all["day"] <= last_data_day]
            .groupby("region")["revenue_gbp"].sum().reset_index()
        )
        reg_pri_map = reg_pri_comp.set_index("region")["revenue_gbp"]
        reg_pct_text = [
            f"{(r - reg_pri_map.get(rgn, 0)) / reg_pri_map.get(rgn, 1) * 100:+.0f}%"
            if reg_pri_map.get(rgn, 0) > 0 else ""
            for rgn, r in zip(reg_cur["region"], reg_cur["revenue_gbp"])
        ]
        fig_r = go.Figure()
        fig_r.add_trace(go.Bar(
            x=reg_pri_comp["region"], y=reg_pri_comp["revenue_gbp"],
            name=f"{prior} (same days)", marker_color=COLORS["pri"], opacity=0.75,
            customdata=[gbp(v) for v in reg_pri_comp["revenue_gbp"]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_r.add_trace(go.Bar(
            x=reg_cur["region"], y=reg_cur["revenue_gbp"],
            name=str(year), marker_color=COLORS["cur"],
            text=reg_pct_text, textposition="outside",
            textfont=dict(size=11),
            customdata=[gbp(v) for v in reg_cur["revenue_gbp"]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_r.update_layout(
            barmode="group", yaxis=dict(tickformat="£,.0f", gridcolor="#2a2a3e"),
            height=320, margin=dict(t=10, b=30),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            plot_bgcolor="#0f1117", paper_bgcolor="#0f1117",
            font=dict(color="#ccc"),
        )
        st.plotly_chart(fig_r, use_container_width=True)

    with col_r:
        st.subheader("By Category")
        cat_cur = (cur_df_all.groupby("product_type")["revenue_gbp"]
                   .sum().sort_values(ascending=False).head(12))
        cat_pri_ser = (
            pri_df_all[pri_df_all["day"] <= last_data_day]
            .groupby("product_type")["revenue_gbp"].sum()
        )
        cat_pct_text = [
            f"{(float(cat_cur.get(pt, 0)) - float(cat_pri_ser.get(pt, 0))) / float(cat_pri_ser.get(pt, 1)) * 100:+.0f}%"
            if float(cat_pri_ser.get(pt, 0)) > 0 else ""
            for pt in cat_cur.index[::-1]
        ]
        cat_pri_vals = [float(cat_pri_ser.get(pt, 0)) for pt in cat_cur.index[::-1]]
        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(
            y=cat_cur.index[::-1],
            x=cat_pri_vals,
            name=f"{prior} (same days)", marker_color=COLORS["pri"], opacity=0.75,
            orientation="h",
            customdata=[gbp(v) for v in cat_pri_vals],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_c.add_trace(go.Bar(
            y=cat_cur.index[::-1],
            x=cat_cur.values[::-1],
            name=str(year), marker_color=COLORS["cur"],
            orientation="h",
            text=cat_pct_text, textposition="outside",
            textfont=dict(size=11),
            customdata=[gbp(v) for v in cat_cur.values[::-1]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_c.update_layout(
            barmode="group", xaxis=dict(tickformat="£,.0f", gridcolor="#2a2a3e"),
            height=320, margin=dict(t=10, b=30),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            plot_bgcolor="#0f1117", paper_bgcolor="#0f1117",
            font=dict(color="#ccc"),
        )
        st.plotly_chart(fig_c, use_container_width=True)


# ── 12-MONTH VIEW ─────────────────────────────────────────────────────────────

else:
    cur_by_month = cur_df_all.groupby("month_num")["revenue_gbp"].sum()
    pri_by_month = pri_df_all.groupby("month_num")["revenue_gbp"].sum()

    cur_total = float(cur_by_month.sum())
    pri_total = float(pri_by_month.sum())
    yoy_pct   = (cur_total - pri_total) / pri_total * 100 if pri_total else 0

    # ── Title + KPI cards ─────────────────────────────────────────────────────
    st.title(f"Global Gross Revenue — {year} vs {prior}")

    k1, k2, k3 = st.columns(3)
    k1.metric(f"{year} Revenue (YTD)", gbp(cur_total),
              delta=pct(yoy_pct) + " YoY", delta_color="normal")
    k2.metric(f"{prior} Full Year", gbp(pri_total))
    months_with_data = int(cur_by_month.shape[0])
    k3.metric("Avg monthly revenue", gbp(cur_total / months_with_data) if months_with_data else "—")

    # ── Main line chart ───────────────────────────────────────────────────────
    fig = go.Figure()

    pri_x = [MONTH_SHORT[int(m) - 1] for m in sorted(pri_by_month.index)]
    pri_y = [float(pri_by_month[m]) for m in sorted(pri_by_month.index)]
    fig.add_trace(go.Scatter(
        x=pri_x, y=pri_y,
        mode="lines+markers",
        name=str(prior),
        line=dict(color=COLORS["pri"], dash="dash", width=1.5),
        marker=dict(size=6),
        customdata=[gbp(v) for v in pri_y],
        hovertemplate="%{customdata}<extra></extra>",
    ))

    cur_x = [MONTH_SHORT[int(m) - 1] for m in sorted(cur_by_month.index)]
    cur_y = [float(cur_by_month[m]) for m in sorted(cur_by_month.index)]
    fig.add_trace(go.Scatter(
        x=cur_x, y=cur_y,
        mode="lines+markers",
        name=str(year),
        line=dict(color=COLORS["cur"], width=2),
        marker=dict(size=6),
        customdata=[gbp(v) for v in cur_y],
        hovertemplate="%{customdata}<extra></extra>",
    ))

    # Promo indicators (small coloured triangles below x-axis)
    if show_promos:
        promo_months_cur = load_promo_months(year)
        promo_months_pri = load_promo_months(prior)
        for m in range(1, 13):
            has_cur = m in promo_months_cur
            has_pri = m in promo_months_pri
            if has_cur or has_pri:
                color = COLORS["cur"] if has_cur else COLORS["pri"]
                # Subtle annotation tag below chart
                fig.add_annotation(
                    x=MONTH_SHORT[m - 1],
                    y=0, yref="paper",
                    text="promo",
                    showarrow=False,
                    font=dict(size=8, color=color),
                    yanchor="top",
                    yshift=-4,
                )

    fig.update_layout(**_chart_layout(
        f"Monthly Revenue — {year} vs {prior}",
        xaxis=dict(
            categoryorder="array",
            categoryarray=MONTH_SHORT,
            gridcolor="#2a2a3e",
        ),
    ))
    st.plotly_chart(fig, use_container_width=True)

    # ── Secondary charts ──────────────────────────────────────────────────────
    col_l, col_r = st.columns(2)

    with col_l:
        st.subheader("By Region")
        reg_cur = cur_df_all.groupby("region")["revenue_gbp"].sum().reset_index()
        reg_pri = pri_df_all.groupby("region")["revenue_gbp"].sum().reset_index()
        reg_pri_map = reg_pri.set_index("region")["revenue_gbp"]
        reg_pct_text = [
            f"{(r - reg_pri_map.get(rgn, 0)) / reg_pri_map.get(rgn, 1) * 100:+.0f}%"
            if reg_pri_map.get(rgn, 0) > 0 else ""
            for rgn, r in zip(reg_cur["region"], reg_cur["revenue_gbp"])
        ]
        fig_r = go.Figure()
        fig_r.add_trace(go.Bar(
            x=reg_pri["region"], y=reg_pri["revenue_gbp"],
            name=str(prior), marker_color=COLORS["pri"], opacity=0.75,
            customdata=[gbp(v) for v in reg_pri["revenue_gbp"]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_r.add_trace(go.Bar(
            x=reg_cur["region"], y=reg_cur["revenue_gbp"],
            name=str(year), marker_color=COLORS["cur"],
            text=reg_pct_text, textposition="outside",
            textfont=dict(size=11),
            customdata=[gbp(v) for v in reg_cur["revenue_gbp"]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_r.update_layout(
            barmode="group", yaxis=dict(tickformat="£,.0f", gridcolor="#2a2a3e"),
            height=320, margin=dict(t=10, b=30),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            plot_bgcolor="#0f1117", paper_bgcolor="#0f1117", font=dict(color="#ccc"),
        )
        st.plotly_chart(fig_r, use_container_width=True)

    with col_r:
        st.subheader("By Category")
        cat_cur = (cur_df_all.groupby("product_type")["revenue_gbp"]
                   .sum().sort_values(ascending=False).head(12))
        cat_pri_ser = pri_df_all.groupby("product_type")["revenue_gbp"].sum()
        cat_pct_text = [
            f"{(float(cat_cur.get(pt, 0)) - float(cat_pri_ser.get(pt, 0))) / float(cat_pri_ser.get(pt, 1)) * 100:+.0f}%"
            if float(cat_pri_ser.get(pt, 0)) > 0 else ""
            for pt in cat_cur.index[::-1]
        ]
        cat_pri_vals_ann = [float(cat_pri_ser.get(pt, 0)) for pt in cat_cur.index[::-1]]
        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(
            y=cat_cur.index[::-1],
            x=cat_pri_vals_ann,
            name=str(prior), marker_color=COLORS["pri"], opacity=0.75,
            orientation="h",
            customdata=[gbp(v) for v in cat_pri_vals_ann],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_c.add_trace(go.Bar(
            y=cat_cur.index[::-1],
            x=cat_cur.values[::-1],
            name=str(year), marker_color=COLORS["cur"],
            orientation="h",
            text=cat_pct_text, textposition="outside",
            textfont=dict(size=11),
            customdata=[gbp(v) for v in cat_cur.values[::-1]],
            hovertemplate="%{customdata}<extra></extra>",
        ))
        fig_c.update_layout(
            barmode="group", xaxis=dict(tickformat="£,.0f", gridcolor="#2a2a3e"),
            height=320, margin=dict(t=10, b=30),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            plot_bgcolor="#0f1117", paper_bgcolor="#0f1117", font=dict(color="#ccc"),
        )
        st.plotly_chart(fig_c, use_container_width=True)
